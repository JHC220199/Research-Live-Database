"""
NRLA Data Dashboard — Automated Data Updater
=============================================
Fetches the latest statistics from official UK sources and updates the JSON
data files in /data/. Designed to run daily via GitHub Actions.

Sources:
  - ONS Time Series API  (CPI, CPIH, GDP, earnings, PIPR)
  - Bank of England IADB (base rate, mortgage advances, BTL share)
  - MHCLG Live Table 213 (dwelling completions — England)
  - Stats Wales Data Cube (dwelling completions — Wales)
  - MoJ / HMCTS          (landlord possession statistics)
  - EHS / FRS            (annual survey releases — download link discovery)

The script only writes a JSON file if the data has changed, avoiding
unnecessary Git commits in the GitHub Actions workflow.
"""

import json
import os
import re
import sys
import time
from datetime import datetime, timezone
from pathlib import Path

import requests
from bs4 import BeautifulSoup

# ── Paths ────────────────────────────────────────────────────────────────────
ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "data"

# ── Shared HTTP session ───────────────────────────────────────────────────────
SESSION = requests.Session()
SESSION.headers.update({
    "User-Agent": "NRLA-Data-Dashboard/1.0 (github.com/nrla; data@nrla.org.uk)"
})

# ── Logging helpers ───────────────────────────────────────────────────────────
def log(msg):   print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}")
def warn(msg):  print(f"[{datetime.now().strftime('%H:%M:%S')}] ⚠  {msg}", file=sys.stderr)
def error(msg): print(f"[{datetime.now().strftime('%H:%M:%S')}] ✖  {msg}", file=sys.stderr)


# ═══════════════════════════════════════════════════════════════════════════════
#  DATA FETCHING FUNCTIONS
# ═══════════════════════════════════════════════════════════════════════════════

# ── ONS Time Series API ───────────────────────────────────────────────────────

def fetch_ons_timeseries(cdid: str) -> dict | None:
    """
    Fetch the latest value for an ONS time series CDID code.
    Returns {"value": "...", "period": "..."} or None on failure.

    API docs: https://api.ons.gov.uk/v1/
    """
    url = f"https://api.ons.gov.uk/v1/timeseries/{cdid.lower()}/data"
    try:
        r = SESSION.get(url, timeout=20)
        r.raise_for_status()
        data = r.json()

        # Try months first (most recent for monthly series), then quarters, then years
        for period_type in ("months", "quarters", "years"):
            entries = data.get(period_type, [])
            if entries:
                latest = entries[-1]
                return {
                    "value": latest.get("value"),
                    "period": latest.get("date", latest.get("label", ""))
                }
        warn(f"ONS {cdid}: no period data found in response")
        return None
    except Exception as e:
        error(f"ONS {cdid}: {e}")
        return None


def fetch_ons_pipr_uk() -> dict | None:
    """
    Fetch the latest PIPR (Price Index of Private Rents) annual rate for the UK.
    The PIPR is published in a bulletin; we download the accompanying dataset CSV.
    """
    # The PIPR CSV is published at a stable path on the ONS website
    csv_url = (
        "https://www.ons.gov.uk/generator?format=csv"
        "&uri=/economy/inflationandpriceindices/bulletins/"
        "privaterentalpricegreatbritain/latest"
    )
    # Fallback: scrape the bulletin page for the dataset download link
    bulletin_url = (
        "https://www.ons.gov.uk/economy/inflationandpriceindices/bulletins/"
        "privaterentalpricegreatbritain/latest"
    )
    try:
        # Try to get the dataset directly
        r = SESSION.get(
            "https://www.ons.gov.uk/economy/inflationandpriceindices/"
            "datasets/priceindexofprivaterentspipr/current/pipr.csv",
            timeout=30
        )
        if r.status_code == 200 and r.text.strip():
            lines = [l for l in r.text.splitlines() if l.strip()]
            # Find the UK annual rate row — typically labelled "United Kingdom" or "UK"
            # and the "Annual percentage change" column
            # Parse header to find column positions
            header = None
            for i, line in enumerate(lines):
                if "Period" in line or "period" in line or "Date" in line:
                    header = line.split(",")
                    data_start = i + 1
                    break
            if header:
                # Look for most recent UK annual rate
                for line in reversed(lines[data_start:]):
                    cells = line.split(",")
                    if len(cells) > 1 and cells[0].strip():
                        period = cells[0].strip()
                        # Try to find a numeric value
                        for cell in cells[1:]:
                            try:
                                val = float(cell.strip())
                                return {"value": str(round(val, 1)), "period": period}
                            except (ValueError, AttributeError):
                                continue

        # Fallback: scrape the bulletin page headline figure
        r = SESSION.get(bulletin_url, timeout=20)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "html.parser")
        # Look for the headline statistic (ONS bulletins have a .headline-figures section)
        headline = soup.find(class_=re.compile(r"headline|statistic", re.I))
        if headline:
            text = headline.get_text(strip=True)
            match = re.search(r"([\d.]+)\s*%", text)
            if match:
                return {"value": match.group(1), "period": "Latest"}

        warn("PIPR: could not parse data from CSV or bulletin page")
        return None

    except Exception as e:
        error(f"PIPR: {e}")
        return None


# ── Bank of England IADB API ──────────────────────────────────────────────────

def fetch_boe_series(series_code: str, label: str) -> dict | None:
    """
    Fetch the latest value for a Bank of England statistical series.
    Uses the BoE's IADB CSV export endpoint.
    """
    url = (
        f"https://www.bankofengland.co.uk/boeapps/iadb/fromshowcolumns.asp"
        f"?csv.x=yes&Datefrom=01/Jan/2023&Dateto=now"
        f"&SeriesCodes={series_code}&CSVF=TT&UsingCodes=Y"
    )
    try:
        r = SESSION.get(url, timeout=20)
        r.raise_for_status()
        lines = [l for l in r.text.splitlines() if l.strip()]
        # BoE CSV format: Date,SeriesCode rows (after a header)
        data_lines = []
        for line in lines:
            parts = line.split(",")
            if len(parts) >= 2 and re.match(r"\d{2} \w+ \d{4}", parts[0].strip()):
                data_lines.append(parts)
        if data_lines:
            latest = data_lines[-1]
            date_str = latest[0].strip()
            value = latest[1].strip()
            # Format date to "Mon YYYY"
            try:
                dt = datetime.strptime(date_str, "%d %b %Y")
                period = dt.strftime("%b %Y")
            except ValueError:
                period = date_str
            return {"value": value, "period": period}
        warn(f"BoE {series_code}: no data rows found")
        return None
    except Exception as e:
        error(f"BoE {series_code} ({label}): {e}")
        return None


def fetch_boe_btl_proportion() -> dict | None:
    """
    Fetch the latest buy-to-let proportion from BoE Table 1.33 (MLAR data).
    This is published as an Excel file — we download and parse it.
    """
    # BoE MLAR (Mortgage Lenders and Administrators Return) Excel file
    mlar_page = "https://www.bankofengland.co.uk/statistics/mortgage-lenders-and-administrators"
    try:
        r = SESSION.get(mlar_page, timeout=20)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "html.parser")
        # Find Excel download link
        links = soup.find_all("a", href=re.compile(r"\.xlsx?", re.I))
        for link in links:
            href = link.get("href", "")
            text = link.get_text(strip=True).lower()
            if "table" in text or "mlar" in text or "1.33" in text:
                if not href.startswith("http"):
                    href = "https://www.bankofengland.co.uk" + href
                try:
                    import io
                    import openpyxl
                    xr = SESSION.get(href, timeout=30)
                    xr.raise_for_status()
                    wb = openpyxl.load_workbook(io.BytesIO(xr.content), data_only=True)
                    # Table 1.33 sheet
                    for sheetname in wb.sheetnames:
                        if "1.33" in sheetname or "btl" in sheetname.lower():
                            ws = wb[sheetname]
                            rows = list(ws.iter_rows(values_only=True))
                            # Find the last row with a numeric value
                            for row in reversed(rows):
                                period_val = row[0]
                                btl_val = None
                                for cell in row[1:]:
                                    if isinstance(cell, (int, float)):
                                        btl_val = cell
                                        break
                                if period_val and btl_val is not None:
                                    return {
                                        "value": str(round(float(btl_val), 1)),
                                        "period": str(period_val)
                                    }
                except Exception as xe:
                    warn(f"BoE BTL Excel parse failed: {xe}")
        warn("BoE BTL: could not find Table 1.33 Excel file")
        return None
    except Exception as e:
        error(f"BoE BTL proportion: {e}")
        return None


# ── MHCLG Live Table 213 ─────────────────────────────────────────────────────

def fetch_mhclg_table213() -> dict | None:
    """
    Download MHCLG Live Table 213 (permanent dwellings completed by tenure, England).
    Returns a dict of tenure keys -> {value, period}.
    """
    page_url = "https://www.gov.uk/government/statistical-data-sets/live-tables-on-house-building"
    try:
        import io
        import openpyxl

        r = SESSION.get(page_url, timeout=20)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "html.parser")

        # Find the Table 213 download link
        file_url = None
        for link in soup.find_all("a", href=True):
            href = link["href"]
            text = link.get_text(strip=True)
            if re.search(r"table.{0,5}213", text, re.I) or \
               re.search(r"table.{0,5}213", href, re.I):
                if re.search(r"\.(xlsx?|csv)$", href, re.I):
                    file_url = href if href.startswith("http") else "https://www.gov.uk" + href
                    break

        if not file_url:
            # Try broader search for housebuilding Excel
            for link in soup.find_all("a", href=re.compile(r"table-213|housebuilding.*213", re.I)):
                file_url = link["href"]
                if not file_url.startswith("http"):
                    file_url = "https://www.gov.uk" + file_url
                break

        if not file_url:
            warn("MHCLG Table 213: download link not found")
            return None

        log(f"MHCLG Table 213: downloading from {file_url}")
        xr = SESSION.get(file_url, timeout=60)
        xr.raise_for_status()

        wb = openpyxl.load_workbook(io.BytesIO(xr.content), data_only=True)

        # The worksheet is typically named "213" or similar
        ws = None
        for name in wb.sheetnames:
            if "213" in name:
                ws = wb[name]
                break
        if ws is None:
            ws = wb.active

        # Parse the table: look for tenure columns (Private, HA, LA, Total)
        # and find the most recent quarterly row
        rows = list(ws.iter_rows(values_only=True))

        # Find header row (contains "Quarter" or "Year" and tenure labels)
        header_idx = None
        col_map = {}
        for i, row in enumerate(rows):
            row_text = " ".join(str(c) for c in row if c)
            if re.search(r"private|housing assoc|local auth|total", row_text, re.I):
                header_idx = i
                for j, cell in enumerate(row):
                    cell_str = str(cell).lower() if cell else ""
                    if "private" in cell_str: col_map["private"] = j
                    elif "housing assoc" in cell_str or "rsl" in cell_str: col_map["ha"] = j
                    elif "local auth" in cell_str: col_map["la"] = j
                    elif "total" in cell_str: col_map["total"] = j
                break

        if not header_idx or not col_map:
            warn("MHCLG Table 213: could not parse headers")
            return None

        # Find the last data row with a period (quarterly format: "2025 Q3")
        result = {}
        for row in reversed(rows[header_idx + 1:]):
            period_cell = row[0]
            if not period_cell:
                continue
            period_str = str(period_cell).strip()
            if not re.search(r"\d{4}", period_str):
                continue
            for key, col_idx in col_map.items():
                if col_idx < len(row) and row[col_idx] is not None:
                    try:
                        result[key] = {
                            "value": str(int(row[col_idx])),
                            "period": period_str
                        }
                    except (ValueError, TypeError):
                        pass
            if result:
                break

        return result if result else None

    except Exception as e:
        error(f"MHCLG Table 213: {e}")
        return None


# ── Stats Wales ───────────────────────────────────────────────────────────────

def fetch_stats_wales_completions() -> dict | None:
    """
    Fetch new dwelling completions from Stats Wales.
    Uses the Stats Wales OData API.
    """
    # Stats Wales API for new house building
    # Dataset: "new-dwellings-completed-by-local-authority-tenure-dwelling-type-and-number-of-bedrooms"
    api_url = (
        "https://statswales.gov.wales/api/v1/dataset/"
        "hous0302/data?$top=5&$orderby=Year_ItemName_ENG desc"
    )
    try:
        r = SESSION.get(api_url, timeout=20)
        if r.status_code == 200:
            data = r.json()
            # Parse the response for total completions
            rows = data.get("value", [])
            if rows:
                latest = rows[0]
                return {
                    "total": {
                        "value": str(latest.get("Data", "")),
                        "period": latest.get("Year_ItemName_ENG", "")
                    }
                }
    except Exception:
        pass

    # Fallback: scrape the Stats Wales page
    try:
        fallback_url = "https://statswales.gov.wales/Catalogue/Housing/New-House-Building"
        r = SESSION.get(fallback_url, timeout=20)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "html.parser")
        # Look for the most recent year in a data table
        tables = soup.find_all("table")
        for table in tables:
            rows = table.find_all("tr")
            for row in reversed(rows):
                cells = [td.get_text(strip=True) for td in row.find_all(["td", "th"])]
                if len(cells) >= 2 and re.search(r"\d{4}", cells[0]):
                    try:
                        val = int(cells[-1].replace(",", ""))
                        return {"total": {"value": str(val), "period": cells[0]}}
                    except (ValueError, IndexError):
                        continue
    except Exception as e:
        warn(f"Stats Wales fallback: {e}")

    return None


# ── MoJ Landlord Possession Statistics ───────────────────────────────────────

def fetch_moj_possession() -> dict | None:
    """
    Fetch landlord possession statistics from MoJ / HMCTS.
    Downloads the latest Excel file from the collection page.
    """
    collection_url = (
        "https://www.gov.uk/government/collections/"
        "mortgage-and-landlord-possession-statistics"
    )
    try:
        import io
        import openpyxl

        r = SESSION.get(collection_url, timeout=20)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "html.parser")

        # Find the most recent publication link, then its Excel file
        # Publications are listed with dates — find the most recent
        pub_links = soup.find_all("a", href=re.compile(
            r"/government/statistics/.*possession", re.I
        ))

        file_url = None
        for pub_link in pub_links[:3]:  # Try the first 3 (most recent)
            pub_href = pub_link["href"]
            if not pub_href.startswith("http"):
                pub_href = "https://www.gov.uk" + pub_href
            try:
                pr = SESSION.get(pub_href, timeout=15)
                pr.raise_for_status()
                psoup = BeautifulSoup(pr.text, "html.parser")
                for link in psoup.find_all("a", href=re.compile(r"\.xlsx?", re.I)):
                    href = link["href"]
                    link_text = link.get_text(strip=True).lower()
                    # We want the main tables file (not the user guide)
                    if "guide" not in link_text and "technical" not in link_text:
                        file_url = href if href.startswith("http") else "https://www.gov.uk" + href
                        break
                if file_url:
                    break
            except Exception:
                continue

        if not file_url:
            warn("MoJ possession: Excel download link not found")
            return None

        log(f"MoJ possession: downloading from {file_url}")
        xr = SESSION.get(file_url, timeout=60)
        xr.raise_for_status()

        wb = openpyxl.load_workbook(io.BytesIO(xr.content), data_only=True)
        result = {}

        # Table 4: Claims issued and bailiff repossessions
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            sheet_lower = sheet_name.lower()

            if "table 4" in sheet_lower or "tab4" in sheet_lower:
                # Find the last row with quarterly data (format: "YYYY Q1" etc.)
                for row in reversed(rows):
                    if not row[0]:
                        continue
                    period_str = str(row[0]).strip()
                    if re.match(r"\d{4}\s*Q[1-4]", period_str):
                        # Column order typically: Quarter, Claims, Orders, Warrants, Repossessions
                        numeric_cells = [
                            c for c in row[1:] if isinstance(c, (int, float))
                        ]
                        if len(numeric_cells) >= 2:
                            result["claims_issued"] = {
                                "value": str(int(numeric_cells[0])),
                                "period": period_str
                            }
                            result["repossessions_bailiffs"] = {
                                "value": str(int(numeric_cells[-1])),
                                "period": period_str
                            }
                            break

            elif "table 6" in sheet_lower or "tab6" in sheet_lower:
                # Mean and median time
                for row in reversed(rows):
                    if not row[0]:
                        continue
                    period_str = str(row[0]).strip()
                    if re.match(r"\d{4}\s*Q[1-4]", period_str):
                        numeric_cells = [
                            c for c in row[1:] if isinstance(c, (int, float))
                        ]
                        if len(numeric_cells) >= 2:
                            result["mean_time_all"] = {
                                "value": str(round(float(numeric_cells[0]), 1)),
                                "period": period_str
                            }
                            result["median_time_all"] = {
                                "value": str(round(float(numeric_cells[1]), 1)),
                                "period": period_str
                            }
                            break

            elif "table 7" in sheet_lower or "tab7" in sheet_lower:
                for row in reversed(rows):
                    if not row[0]:
                        continue
                    period_str = str(row[0]).strip()
                    if re.match(r"\d{4}\s*Q[1-4]", period_str):
                        numeric_cells = [
                            c for c in row[1:] if isinstance(c, (int, float))
                        ]
                        if numeric_cells:
                            result["claims_prs"] = {
                                "value": str(int(numeric_cells[0])),
                                "period": period_str
                            }
                            break

        return result if result else None

    except Exception as e:
        error(f"MoJ possession: {e}")
        return None


# ── Annual Release Finder (EHS / FRS) ─────────────────────────────────────────

def find_latest_annual_release(collection_url: str, publication_pattern: str,
                               label: str) -> tuple[str | None, str | None]:
    """
    Scrapes a gov.uk collection page to find the most recent annual release.
    Returns (publication_url, year_string) or (None, None).
    """
    try:
        r = SESSION.get(collection_url, timeout=15)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "html.parser")

        year_pattern = re.compile(r"20\d\d")
        pub_links = soup.find_all("a", href=re.compile(publication_pattern, re.I))

        best_link = None
        best_year = 0
        for link in pub_links:
            years = year_pattern.findall(link.get("href", "") + link.get_text())
            for y in years:
                if int(y) > best_year:
                    best_year = int(y)
                    best_link = link

        if best_link:
            href = best_link["href"]
            if not href.startswith("http"):
                href = "https://www.gov.uk" + href
            log(f"{label}: latest release found at {href} ({best_year})")
            return href, str(best_year)

        warn(f"{label}: no release found at {collection_url}")
        return None, None

    except Exception as e:
        error(f"{label} collection page: {e}")
        return None, None


# ═══════════════════════════════════════════════════════════════════════════════
#  JSON UPDATE FUNCTIONS
# ═══════════════════════════════════════════════════════════════════════════════

def load_json(path: Path) -> dict:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

def save_json(path: Path, data: dict) -> bool:
    """Write JSON, return True if content changed."""
    new_content = json.dumps(data, indent=2, ensure_ascii=False)
    if path.exists():
        old_content = path.read_text(encoding="utf-8")
        if old_content == new_content:
            return False
    path.write_text(new_content, encoding="utf-8")
    return True

def now_iso() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

def update_dataset(datasets: list, dataset_id: str, fetched: dict | None) -> bool:
    """Update a dataset entry in-place. Returns True if changed."""
    for ds in datasets:
        if ds["id"] == dataset_id:
            if fetched and fetched.get("value") is not None:
                old = ds.get("latest", {})
                new_val = {"value": str(fetched["value"]),
                           "period": fetched.get("period", old.get("period")),
                           "fetched_at": now_iso()}
                if old != new_val:
                    ds["latest"] = new_val
                    return True
    return False


# ═══════════════════════════════════════════════════════════════════════════════
#  SECTION UPDATE RUNNERS
# ═══════════════════════════════════════════════════════════════════════════════

def update_macro():
    log("=== Updating Macro-Economic data ===")
    path = DATA_DIR / "macro.json"
    data = load_json(path)
    ds = data["datasets"]
    changed = False

    ons_cdid_map = {
        "cpi_annual_rate":   "D7G7",
        "cpi_index":         "D7BT",
        "cpih_annual_rate":  "L55O",
        "cpih_index":        "L522",
        "gdp_qoq":           "IHYQ",
        "gdp_yoy":           "IHYR",
        "awe_total_pay":     "K54U",
        "real_earnings_index":  "A2FD",
        "real_earnings_growth": "A3WV",
    }

    for dataset_id, cdid in ons_cdid_map.items():
        log(f"  ONS {cdid} → {dataset_id}")
        result = fetch_ons_timeseries(cdid)
        if update_dataset(ds, dataset_id, result):
            log(f"    Updated: {result}")
            changed = True
        else:
            log(f"    No change")
        time.sleep(0.5)  # Be polite to the API

    # PIPR (private rents)
    log("  PIPR annual rate")
    pipr = fetch_ons_pipr_uk()
    if update_dataset(ds, "pipr_annual_rate_uk", pipr):
        log(f"    Updated: {pipr}")
        changed = True
    # PIPR index — same bulletin, typically published together
    if update_dataset(ds, "pipr_index_uk", pipr):
        changed = True

    # Bank of England base rate
    log("  BoE base rate (IUMABEDR)")
    boe_rate = fetch_boe_series("IUMABEDR", "base rate")
    if update_dataset(ds, "boe_base_rate", boe_rate):
        log(f"    Updated: {boe_rate}")
        changed = True

    # BoE gross advances — series LPMB3R6 (Table 1.21, total gross advances)
    log("  BoE gross advances (LPMB3R6)")
    boe_adv = fetch_boe_series("LPMB3R6", "gross advances")
    if update_dataset(ds, "boe_gross_advances", boe_adv):
        log(f"    Updated: {boe_adv}")
        changed = True

    # BoE BTL proportion
    log("  BoE BTL proportion (Table 1.33)")
    boe_btl = fetch_boe_btl_proportion()
    if update_dataset(ds, "boe_btl_proportion", boe_btl):
        log(f"    Updated: {boe_btl}")
        changed = True

    if changed:
        data["last_fetched"] = now_iso()
        save_json(path, data)
        log("  macro.json saved ✓")
    else:
        log("  No changes to macro.json")


def update_housing_tenure():
    log("=== Updating Housing Trends (FRS) ===")
    path = DATA_DIR / "housing_tenure.json"
    data = load_json(path)

    # FRS is published annually. We note the expected period but can't
    # auto-extract individual cell values easily without knowing the exact
    # Excel layout which changes each year.
    # We discover whether a new release exists and update the period/fetched_at.

    collection_url = "https://www.gov.uk/government/collections/family-resources-survey--2"
    pub_url, year = find_latest_annual_release(
        collection_url,
        r"family-resources-survey",
        "FRS"
    )

    if not year:
        log("  FRS: no new release detected")
        return

    changed = False
    for ds in data["datasets"]:
        if ds["status"] == "active":
            latest = ds.get("latest", {})
            # Check if the year in the stored period differs from the newly found one
            stored_period = latest.get("period", "")
            if year not in stored_period:
                ds["latest"] = {
                    "value": latest.get("value"),  # preserve last known value
                    "period": f"{int(year)-1}/{str(year)[2:]}",  # e.g. "2023/24"
                    "fetched_at": now_iso()
                }
                changed = True

    if changed:
        data["last_fetched"] = now_iso()
        save_json(path, data)
        log("  housing_tenure.json updated with new release period ✓")
    else:
        log("  No changes to housing_tenure.json")


def update_dwellings():
    log("=== Updating Dwellings data ===")
    path = DATA_DIR / "dwellings.json"
    data = load_json(path)
    ds = data["datasets"]
    changed = False

    # MHCLG Table 213
    log("  MHCLG Table 213 (England completions)")
    mhclg = fetch_mhclg_table213()
    if mhclg:
        for key, dst_id in [
            ("private", "mhclg_completions_private"),
            ("ha",      "mhclg_completions_ha"),
            ("la",      "mhclg_completions_la"),
            ("total",   "mhclg_completions_total"),
        ]:
            if key in mhclg and update_dataset(ds, dst_id, mhclg[key]):
                log(f"    {dst_id}: {mhclg[key]}")
                changed = True

    # Stats Wales
    log("  Stats Wales (Wales completions)")
    wales = fetch_stats_wales_completions()
    if wales:
        if "total" in wales and update_dataset(ds, "stats_wales_completions_total", wales["total"]):
            log(f"    stats_wales_completions_total: {wales['total']}")
            changed = True
        if "private" in wales and update_dataset(ds, "stats_wales_completions_private", wales["private"]):
            log(f"    stats_wales_completions_private: {wales['private']}")
            changed = True

    if changed:
        data["last_fetched"] = now_iso()
        save_json(path, data)
        log("  dwellings.json saved ✓")
    else:
        log("  No changes to dwellings.json")


def update_possession():
    log("=== Updating Claims & Bailiffs (MoJ) ===")
    path = DATA_DIR / "possession.json"
    data = load_json(path)
    ds = data["datasets"]

    moj = fetch_moj_possession()
    if not moj:
        log("  No MoJ data retrieved")
        return

    changed = False
    mapping = {
        "claims_issued":          "mlp_claims_issued",
        "repossessions_bailiffs": "mlp_repossessions_bailiffs",
        "claims_prs":             "mlp_claims_prs",
        "mean_time_all":          "mlp_mean_time_all",
        "median_time_all":        "mlp_median_time_all",
    }
    for moj_key, ds_id in mapping.items():
        if moj_key in moj and update_dataset(ds, ds_id, moj[moj_key]):
            log(f"  {ds_id}: {moj[moj_key]}")
            changed = True

    if changed:
        data["last_fetched"] = now_iso()
        save_json(path, data)
        log("  possession.json saved ✓")
    else:
        log("  No changes to possession.json")


def update_ehs():
    log("=== Updating English Housing Survey ===")
    path = DATA_DIR / "ehs.json"
    data = load_json(path)

    # EHS is annual — detect if a new release is available
    collection_url = "https://www.gov.uk/government/collections/english-housing-survey"
    pub_url, year = find_latest_annual_release(
        collection_url,
        r"english-housing-survey",
        "EHS"
    )

    if not year:
        log("  EHS: no new release detected")
        return

    changed = False
    for ds in data["datasets"]:
        if ds["status"] == "active":
            latest = ds.get("latest", {})
            stored_period = latest.get("period", "")
            if year not in stored_period:
                ds["latest"] = {
                    "value": latest.get("value"),
                    "period": year,
                    "fetched_at": now_iso()
                }
                changed = True

    if changed:
        data["last_fetched"] = now_iso()
        save_json(path, data)
        log("  ehs.json updated with new release year ✓")
    else:
        log("  No changes to ehs.json")


# ═══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    log(f"NRLA Data Dashboard updater starting — {datetime.now().strftime('%Y-%m-%d %H:%M UTC')}")
    log(f"Data directory: {DATA_DIR}")

    if not DATA_DIR.exists():
        error(f"Data directory not found: {DATA_DIR}")
        sys.exit(1)

    errors = []

    for label, fn in [
        ("Macro-Economic",   update_macro),
        ("Housing Tenure",   update_housing_tenure),
        ("Dwellings",        update_dwellings),
        ("Claims/Bailiffs",  update_possession),
        ("EHS",              update_ehs),
    ]:
        try:
            fn()
        except Exception as e:
            err_msg = f"{label}: unexpected error — {e}"
            error(err_msg)
            errors.append(err_msg)

    if errors:
        log(f"\nCompleted with {len(errors)} error(s):")
        for e in errors:
            log(f"  ✖ {e}")
        sys.exit(1)
    else:
        log("\nAll sections updated successfully ✓")


if __name__ == "__main__":
    main()
