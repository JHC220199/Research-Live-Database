"""
NRLA Data Dashboard — Automated Data Updater
=============================================
Fetches the latest statistics from official UK sources and updates the JSON
data files in /data/. Designed to run daily via GitHub Actions.
 
Sources:
  - ONS website JSON endpoint  (CPI, CPIH, GDP, earnings, PIPR)
  - Bank of England IADB       (base rate, mortgage advances)
  - MHCLG Live Table 213       (dwelling completions — England)
  - Stats Wales                (dwelling completions — Wales)
  - MoJ / HMCTS CSV zip        (landlord possession statistics)
  - EHS / FRS collection pages (annual survey release detection)
"""
 
import csv
import io
import json
import os
import re
import sys
import time
import zipfile
from datetime import datetime, timezone
from pathlib import Path
 
import requests
import openpyxl
import pandas as pd
from bs4 import BeautifulSoup
 
# ── Paths ─────────────────────────────────────────────────────────────────────
ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "data"
 
# ── Shared HTTP session ───────────────────────────────────────────────────────
SESSION = requests.Session()
SESSION.headers.update({
    "User-Agent": "NRLA-Data-Dashboard/1.0 (github.com/nrla; data@nrla.org.uk)"
})
 
# ── Logging helpers ───────────────────────────────────────────────────────────
def log(msg):   print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}")
def warn(msg):  print(f"[{datetime.now().strftime('%H:%M:%S')}] ⚠  {msg}", file=sys.stderr)
def error(msg): print(f"[{datetime.now().strftime('%H:%M:%S')}] ✖  {msg}", file=sys.stderr)
 
 
# ═══════════════════════════════════════════════════════════════════════════════
#  ONS TIME SERIES
# ═══════════════════════════════════════════════════════════════════════════════
 
# Full ONS website paths for each CDID code.
# The data endpoint is: https://www.ons.gov.uk/{path}/data  → returns JSON
ONS_PATHS = {
    "D7G7": "economy/inflationandpriceindices/timeseries/d7g7",
    "D7BT": "economy/inflationandpriceindices/timeseries/d7bt",
    "L55O": "economy/inflationandpriceindices/timeseries/l55o",
    "L522": "economy/inflationandpriceindices/timeseries/l522",
    "IHYQ": "economy/grossdomesticproductgdp/timeseries/ihyq",
    "IHYR": "economy/grossdomesticproductgdp/timeseries/ihyr",
    "K54U": "employmentandlabourmarket/peopleinwork/earningsandworkinghours/timeseries/k54u",
    "A2FD": "employmentandlabourmarket/peopleinwork/earningsandworkinghours/timeseries/a2fd",
    "A3WV": "employmentandlabourmarket/peopleinwork/earningsandworkinghours/timeseries/a3wv",
    # PIPR — Price Index of Private Rents (standalone dataset, CDIDs not available via MM23)
    # NOTE: PIPR is fetched via fetch_ons_pipr_uk() which downloads the published data file.
}
 
 
def fetch_ons_timeseries(cdid: str) -> dict | None:
    """
    Fetch the latest value for an ONS time series using the website's JSON endpoint.
    Returns {"value": "...", "period": "..."} or None on failure.
    """
    path = ONS_PATHS.get(cdid.upper())
    if not path:
        warn(f"ONS {cdid}: no path configured")
        return None
 
    url = f"https://www.ons.gov.uk/{path}/data"
    try:
        r = SESSION.get(url, timeout=20)
        r.raise_for_status()
        data = r.json()
 
        # Try months first (most recent for monthly series), then quarters, then years
        for period_type in ("months", "quarters", "years"):
            entries = data.get(period_type, [])
            if entries:
                latest = entries[-1]
                raw_date = latest.get("date", latest.get("label", ""))
                # Format monthly dates: "2026 JAN" → "Jan 2026"
                formatted = _format_ons_date(raw_date)
                return {"value": latest.get("value"), "period": formatted}
 
        warn(f"ONS {cdid}: no period data in response")
        return None
 
    except Exception as e:
        error(f"ONS {cdid}: {e}")
        return None
 
 
def _format_ons_date(raw: str) -> str:
    """Convert ONS date strings like '2026 JAN' or '2025 Q3' to readable form."""
    raw = raw.strip()
    # Monthly: "2026 JAN"
    m = re.match(r"(\d{4})\s+([A-Z]{3})$", raw)
    if m:
        try:
            dt = datetime.strptime(f"{m.group(2)} {m.group(1)}", "%b %Y")
            return dt.strftime("%b %Y")
        except ValueError:
            pass
    # Quarterly: "2025 Q3"
    m = re.match(r"(\d{4})\s+(Q[1-4])$", raw)
    if m:
        return f"{m.group(2)} {m.group(1)}"
    return raw
 
 

PIPR_DATASET = ("https://www.ons.gov.uk/economy/inflationandpriceindices/datasets/"
                "priceindexofprivaterentsukmonthlypricestatistics")


def fetch_ons_pipr_uk() -> dict | None:
    """
    PIPR (Price Index of Private Rents) UK annual rate and index.

    Downloads the latest "monthly price statistics" xlsx from the ONS dataset
    page (discovered via the /data JSON endpoint) and reads the last row of the
    United Kingdom block (area code K02000001) in Table 1.

    This replaces the old bulletin HTML scrape — the bulletin endpoint returns
    502 intermittently, and the regex extraction once stored an index value
    (140) as the annual rate. A plausibility window (0.5–25%) guards the rate.

    Returns {"annual_rate": {...}, "index": {...}} or None.
    """
    try:
        r = SESSION.get(PIPR_DATASET + "/data", timeout=30)
        r.raise_for_status()
        editions = r.json().get("datasets", [])
        if not editions:
            warn("PIPR: no editions listed on dataset page")
            return None
        latest_uri = editions[0]["uri"]
        er = SESSION.get(f"https://www.ons.gov.uk{latest_uri}/data", timeout=30)
        er.raise_for_status()
        dls = er.json().get("downloads", [])
        if not dls:
            warn("PIPR: no downloads on latest edition")
            return None
        file_url = f"https://www.ons.gov.uk/file?uri={latest_uri}/{dls[0]['file']}"
        log(f"  PIPR: downloading {dls[0]['file']} ({latest_uri.split('/')[-1]})")
        fr = SESSION.get(file_url, timeout=180)
        fr.raise_for_status()
        wb = openpyxl.load_workbook(io.BytesIO(fr.content), read_only=True, data_only=True)
        ws = wb["Table 1"]
        last_uk = None
        for row in ws.iter_rows(values_only=True):
            if row and len(row) > 6 and str(row[1]).strip() == "K02000001":
                last_uk = row
            elif last_uk is not None and row and str(row[1]).strip() != "K02000001":
                break  # UK block sits at the top of the sheet
        if not last_uk:
            warn("PIPR: UK rows not found in Table 1")
            return None
        period_raw, index_v, annual_v = last_uk[0], _num(last_uk[4]), _num(last_uk[6])
        if isinstance(period_raw, datetime):
            period = period_raw.strftime("%b %Y")
        else:
            try:
                period = datetime.fromisoformat(str(period_raw)[:10]).strftime("%b %Y")
            except ValueError:
                period = str(period_raw)
        out = {}
        if annual_v is not None and 0.5 <= abs(annual_v) <= 25:
            out["annual_rate"] = {"value": _fmt(annual_v, 1), "period": period}
        else:
            warn(f"PIPR: annual rate {annual_v} failed plausibility check")
        if index_v is not None and 50 <= index_v <= 250:
            out["index"] = {"value": _fmt(index_v, 1), "period": period}
        return out or None
    except Exception as e:
        error(f"PIPR: {e}")
        return None


def fetch_boe_series(series_code: str, label: str) -> dict | None:
    """
    Fetch the latest value for a Bank of England statistical series via IADB CSV.
    Handles both monthly and quarterly date formats.
    """
    url = (
        f"https://www.bankofengland.co.uk/boeapps/iadb/fromshowcolumns.asp"
        f"?csv.x=yes&Datefrom=01/Jan/2022&Dateto=now"
        f"&SeriesCodes={series_code}&CSVF=TT&UsingCodes=Y"
    )
    try:
        r = SESSION.get(url, timeout=25)
        r.raise_for_status()
        lines = [l.strip() for l in r.text.splitlines() if l.strip()]
 
        data_lines = []
        for line in lines:
            parts = [p.strip() for p in line.split(",")]
            if len(parts) >= 2:
                # Match both "01 Jan 2024" (monthly) and "01 Jan 2024" style dates
                if re.match(r"\d{2}\s+\w+\s+\d{4}", parts[0]):
                    data_lines.append(parts)
 
        if not data_lines:
            warn(f"BoE {series_code}: no data rows found")
            return None
 
        latest = data_lines[-1]
        date_str = latest[0]
        value = latest[1] if len(latest) > 1 else None
 
        if not value or value in ("", ".", ".."):
            # Sometimes last row is incomplete — try second-to-last
            if len(data_lines) > 1:
                latest = data_lines[-2]
                date_str = latest[0]
                value = latest[1] if len(latest) > 1 else None
 
        try:
            dt = datetime.strptime(date_str, "%d %b %Y")
            period = dt.strftime("%b %Y")
        except ValueError:
            period = date_str
 
        return {"value": value, "period": period}
 
    except Exception as e:
        error(f"BoE {series_code} ({label}): {e}")
        return None
 
 
# ═══════════════════════════════════════════════════════════════════════════════
#  MHCLG LIVE TABLE 213
# ═══════════════════════════════════════════════════════════════════════════════
 
def fetch_mhclg_table213() -> dict | None:
    """
    Download MHCLG Live Table 213 (permanent dwellings completed, England).
    Scrapes the gov.uk page to find the current Excel download URL.
    """
    page_url = "https://www.gov.uk/government/statistical-data-sets/live-tables-on-house-building"
    try:
        import openpyxl
 
        r = SESSION.get(page_url, timeout=20)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "html.parser")
 
        # Find the Table 213 Excel download link
        file_url = None
        for link in soup.find_all("a", href=True):
            href = link["href"]
            text = link.get_text(strip=True)
            if (re.search(r"table.{0,5}213", text, re.I) or
                    re.search(r"table.{0,5}213|LiveTable213", href, re.I)):
                if re.search(r"\.(xlsx?)$", href, re.I):
                    file_url = href if href.startswith("http") else "https://www.gov.uk" + href
                    break
 
        if not file_url:
            # Broader search in assets URLs
            for link in soup.find_all("a", href=re.compile(r"LiveTable213|table213|Table213", re.I)):
                file_url = link["href"]
                if not file_url.startswith("http"):
                    file_url = "https://www.gov.uk" + file_url
                break
 
        if not file_url:
            warn("MHCLG Table 213: download link not found on page")
            return None
 
        log(f"  MHCLG Table 213: {file_url}")
        xr = SESSION.get(file_url, timeout=60)
        xr.raise_for_status()
 
        wb = openpyxl.load_workbook(io.BytesIO(xr.content), data_only=True)
 
        # Find the "213" worksheet
        ws = None
        for name in wb.sheetnames:
            if "213" in name:
                ws = wb[name]
                break
        if ws is None:
            ws = wb.active
 
        rows = list(ws.iter_rows(values_only=True))
 
        # Find header row containing tenure column labels.
        # Table 213 has columns: Period | ...Starts... | All Starts | ...Completions... | All Completions | Notes
        # We want the COMPLETIONS columns only (they come after the Starts block).
        col_map = {}
        header_idx = None
        for i, row in enumerate(rows):
            row_text = " ".join(str(c).lower() for c in row if c)
            if "private" in row_text and "completion" in row_text:
                header_idx = i
                for j, cell in enumerate(row):
                    if not cell:
                        continue
                    s = str(cell).lower()
                    # Skip any starts columns — we only want completions
                    if "start" in s:
                        continue
                    if "private" in s and "enterprise" in s and "completion" in s:
                        col_map["private"] = j
                    elif ("housing assoc" in s or "registered" in s or " rsl" in s) and "completion" in s:
                        col_map["ha"] = j
                    elif "local auth" in s and "completion" in s:
                        col_map["la"] = j
                    elif ("all completion" in s or "total completion" in s or s.strip() == "total"):
                        col_map["total"] = j
                if col_map:
                    break
 
        if not header_idx or not col_map:
            warn("MHCLG Table 213: could not identify header columns")
            return None
 
        # Find the last quarterly data row
        result = {}
        for row in reversed(rows[header_idx + 1:]):
            if not row[0]:
                continue
            period_str = str(row[0]).strip()
            # Quarterly format: "2025 Q3" or "2025Q3" or "Q3 2025"
            if not re.search(r"Q[1-4]", period_str, re.I):
                continue
            for key, col_idx in col_map.items():
                if col_idx < len(row) and row[col_idx] is not None:
                    try:
                        result[key] = {
                            "value": str(int(float(str(row[col_idx]).replace(",", "")))),
                            "period": period_str
                        }
                    except (ValueError, TypeError):
                        pass
            if result:
                break
 
        return result if result else None
 
    except Exception as e:
        error(f"MHCLG Table 213: {e}")
        return None
 
 
# ═══════════════════════════════════════════════════════════════════════════════
#  STATS WALES
# ═══════════════════════════════════════════════════════════════════════════════
 

SW_API = "https://api.stats.gov.wales/v1"
SW_TOPIC_NEW_HOUSE_BUILDING = 66


def fetch_stats_wales_completions() -> dict | None:
    """Wales new dwellings completed, latest quarter: total + private enterprise.
    Uses the StatsWales open data API launched Sept 2025."""
    try:
        h = {"Accept": "application/json"}  # new StatsWales open data service (Sept 2025)
        # 1. find the completions dataset under topic 66
        t = SESSION.get(f"{SW_API}/topic/{SW_TOPIC_NEW_HOUSE_BUILDING}",
                        params={"lang": "en-GB"}, headers=h, timeout=30)
        t.raise_for_status()
        ds_id = None
        for d in t.json().get("datasets", {}).get("data", []):
            if "completed" in d.get("title", "").lower():
                ds_id = d["id"]
                break
        if not ds_id:
            warn("StatsWales: completions dataset not found in topic 66")
            return None
        # 2. latest quarterly period from filters
        f = SESSION.get(f"{SW_API}/{ds_id}/view/filters",
                        params={"lang": "en-GB"}, headers=h, timeout=30)
        f.raise_for_status()
        latest_ref, latest_desc = None, None
        for filt in f.json():
            if filt.get("factTableColumn") == "Period":
                qrefs = []
                def collect(vals):
                    for v in vals:
                        if re.match(r"^\d{6}Q\d$", str(v.get("reference", ""))):
                            qrefs.append((v["reference"], v.get("description", v["reference"])))
                        if v.get("children"):
                            ch = v["children"]
                            if isinstance(ch, list):
                                collect(ch)
                collect(filt.get("values", []))
                if qrefs:
                    latest_ref, latest_desc = max(qrefs, key=lambda x: x[0])
        if not latest_ref:
            warn("StatsWales: no quarterly period found")
            return None
        # 3. pull Wales totals for the latest quarter (JSON filter, reference codes)
        import json as _json
        filt = _json.dumps([
            {"columnName": "LACode", "values": ["600"]},          # Wales
            {"columnName": "Period", "values": [latest_ref]},
            {"columnName": "DwellingType", "values": ["3"]},      # Total
            {"columnName": "Bedroom", "values": ["5"]},           # Total
        ])
        v = SESSION.get(f"{SW_API}/{ds_id}/view", headers=h, timeout=60, params={
            "lang": "en-GB", "page_size": 100, "filter": filt,
        })
        v.raise_for_status()
        body = v.json()
        out = {}
        for row in body.get("data", []):
            cells = [str(x).strip() for x in row]
            if len(cells) < 6:
                continue
            value, measure, period_desc, area, dtype, bed = cells[:6]
            if area != "Wales" or dtype != "Total" or bed != "Total":
                continue
            if period_desc != latest_desc:
                continue
            n = _num(value)
            if n is None:
                continue
            if measure.lower().startswith("total new dwellings completed"):
                out["total"] = _latest(_fmt(n, 0), latest_desc)
            elif measure.lower().startswith("private enterprise"):
                out["private"] = _latest(_fmt(n, 0), latest_desc)
        return out or None
    except Exception as e:
        warn(f"StatsWales API: {e}")
        return None


def fetch_moj_possession() -> dict | None:
    """
    Fetch landlord possession statistics from MoJ / HMCTS.
 
    The MoJ publishes a CSV zip containing court-level data in long format:
        Year, Quarter, possession_type, possession_action, court, region, value
 
    We find the latest publication, download the zip, read the court-level CSV,
    and aggregate national totals by summing across all courts for the latest quarter.
    """
    collection_url = (
        "https://www.gov.uk/government/collections/"
        "mortgage-and-landlord-possession-statistics"
    )
    try:
        r = SESSION.get(collection_url, timeout=20)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "html.parser")
 
        # Find the most recent publication link
        pub_url = None
        for link in soup.find_all("a", href=re.compile(
            r"/government/statistics/mortgage-and-landlord-possession-statistics-(?!earlier|guide)"
        )):
            href = link["href"]
            if not href.startswith("http"):
                href = "https://www.gov.uk" + href
            pub_url = href
            break
 
        if not pub_url:
            warn("MoJ: could not find latest publication link")
            return None
 
        log(f"  MoJ: latest publication → {pub_url}")
 
        # Find the CSV zip on the publication page
        pr = SESSION.get(pub_url, timeout=20)
        pr.raise_for_status()
        psoup = BeautifulSoup(pr.text, "html.parser")
 
        csv_zip_url = None
        for link in psoup.find_all("a", href=re.compile(r"CSVs\.zip", re.I)):
            csv_zip_url = link["href"]
            break
        if not csv_zip_url:
            for link in psoup.find_all("a", href=re.compile(r"\.zip$", re.I)):
                csv_zip_url = link["href"]
                break
 
        if not csv_zip_url:
            warn("MoJ: CSV zip download link not found on publication page")
            return None
 
        log(f"  MoJ: downloading CSV zip from {csv_zip_url}")
        zr = SESSION.get(csv_zip_url, timeout=60)
        zr.raise_for_status()
 
        return _parse_moj_court_csv(zr.content)
 
    except Exception as e:
        error(f"MoJ possession: {e}")
        return None
 
 
def _parse_moj_court_csv(zip_content: bytes) -> dict | None:
    """
    Parse the MoJ court-level CSV (long format) and aggregate national totals.
 
    Expected CSV columns:
        Year, Quarter, possession_type, possession_action, court, region, value
 
    possession_type values: Accelerated_Landlord, Private_Landlord,
                            Social_Landlord, Mortgage, Other
    possession_action values: Claims, Outright_Orders, Suspended_Orders,
                              Warrants, Repossessions, Other
 
    We sum across all courts to get England & Wales national figures.
    """
    try:
        z = zipfile.ZipFile(io.BytesIO(zip_content))
        names = z.namelist()
        log(f"  MoJ zip contents: {names}")
 
        # Find the court-level CSV
        court_csv_name = None
        for name in names:
            if "court" in name.lower() and name.lower().endswith(".csv"):
                court_csv_name = name
                break
 
        if not court_csv_name:
            warn("MoJ: court CSV not found in zip")
            return None
 
        rows = _read_csv_from_zip(z, court_csv_name)
        if not rows:
            return None
 
        # Parse column positions from header
        header = [h.strip().lower() for h in rows[0]]
        def col(name): return next((i for i, h in enumerate(header) if name in h), None)
 
        year_col   = col("year")
        qtr_col    = col("quarter")
        type_col   = col("possession_type") or col("type")
        action_col = col("possession_action") or col("action")
        value_col  = col("value")
 
        if any(c is None for c in [year_col, qtr_col, type_col, action_col, value_col]):
            warn(f"MoJ: unexpected CSV columns: {header}")
            return None
 
        # Find the latest quarter available
        quarters = set()
        for row in rows[1:]:
            if len(row) > qtr_col and re.search(r"Q[1-4]", row[qtr_col], re.I):
                quarters.add((row[year_col], row[qtr_col]))
 
        if not quarters:
            warn("MoJ: no quarterly data found")
            return None
 
        latest_year, latest_qtr = max(quarters, key=lambda x: (x[0], x[1]))
        period = f"{latest_year} {latest_qtr}"
        log(f"  MoJ: latest quarter = {period}")
 
        # Aggregate by summing across all courts for the latest quarter
        totals: dict[tuple, int] = {}
        for row in rows[1:]:
            if len(row) <= value_col:
                continue
            if row[year_col] != latest_year or row[qtr_col] != latest_qtr:
                continue
            key = (row[type_col], row[action_col])
            try:
                totals[key] = totals.get(key, 0) + int(float(row[value_col]))
            except (ValueError, TypeError):
                pass
 
        log(f"  MoJ aggregated {len(totals)} (type, action) combinations")
 
        def get(ptype, action):
            return totals.get((ptype, action), 0)
 
        result = {}
 
        # Total landlord possession claims (all landlord types)
        total_claims = (
            get("Accelerated_Landlord", "Claims") +
            get("Private_Landlord",     "Claims") +
            get("Social_Landlord",      "Claims")
        )
        if total_claims:
            result["claims_issued"] = {"value": str(total_claims), "period": period}
 
        # Total landlord repossessions by bailiff
        total_repos = (
            get("Accelerated_Landlord", "Repossessions") +
            get("Private_Landlord",     "Repossessions") +
            get("Social_Landlord",      "Repossessions")
        )
        if total_repos:
            result["repossessions_bailiffs"] = {"value": str(total_repos), "period": period}
 
        # Private sector claims (private + accelerated landlords)
        prs_claims = (
            get("Private_Landlord",     "Claims") +
            get("Accelerated_Landlord", "Claims")
        )
        if prs_claims:
            result["claims_prs"] = {"value": str(prs_claims), "period": period}
 
        # Accelerated procedure claims only
        acc = get("Accelerated_Landlord", "Claims")
        if acc:
            result["claims_accelerated"] = {"value": str(acc), "period": period}
 
        return result if result else None
 
    except Exception as e:
        error(f"MoJ court CSV parse: {e}")
        return None
 
 
def _read_csv_from_zip(z: zipfile.ZipFile, name: str) -> list[list[str]]:
    """Read a CSV file from a zip archive, returning a list of rows."""
    with z.open(name) as f:
        content = f.read().decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(content))
        return list(reader)
 
 
 
 
# ═══════════════════════════════════════════════════════════════════════════════
#  ANNUAL RELEASE DETECTION (EHS / FRS)
# ═══════════════════════════════════════════════════════════════════════════════
 
def find_latest_annual_release(collection_url: str, publication_pattern: str,
                               label: str) -> tuple[str | None, str | None]:
    """
    Scrape a gov.uk collection page to find the most recent annual release year.
    Returns (publication_url, year_string) or (None, None).
    """
    try:
        r = SESSION.get(collection_url, timeout=15)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "html.parser")
 
        year_pattern = re.compile(r"20\d\d")
        pub_links = soup.find_all("a", href=re.compile(publication_pattern, re.I))
 
        best_link = None
        best_year = 0
        for link in pub_links:
            years = year_pattern.findall(link.get("href", "") + link.get_text())
            for y in years:
                if int(y) > best_year:
                    best_year = int(y)
                    best_link = link
 
        if best_link:
            href = best_link["href"]
            if not href.startswith("http"):
                href = "https://www.gov.uk" + href
            log(f"  {label}: latest release → {href} ({best_year})")
            return href, str(best_year)
 
        warn(f"{label}: no release link found at {collection_url}")
        return None, None
 
    except Exception as e:
        error(f"{label} collection page: {e}")
        return None, None
 
 
# ═══════════════════════════════════════════════════════════════════════════════
#  JSON FILE HELPERS
# ═══════════════════════════════════════════════════════════════════════════════
 
def load_json(path: Path) -> dict:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)
 
 
def save_json(path: Path, data: dict) -> bool:
    """Write JSON. Returns True if content changed."""
    new_content = json.dumps(data, indent=2, ensure_ascii=False)
    if path.exists():
        old_content = path.read_text(encoding="utf-8")
        if old_content == new_content:
            return False
    path.write_text(new_content, encoding="utf-8")
    return True
 
 
def now_iso() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
 
 
def update_dataset(datasets: list, dataset_id: str, fetched: dict | None) -> bool:
    """Update a dataset entry in-place. Returns True if changed."""
    if not fetched or fetched.get("value") is None:
        return False
    for ds in datasets:
        if ds["id"] == dataset_id:
            old = ds.get("latest", {})
            new_val = {
                "value": str(fetched["value"]),
                "period": fetched.get("period", old.get("period")),
                "fetched_at": now_iso()
            }
            if old.get("value") != new_val["value"] or old.get("period") != new_val["period"]:
                ds["latest"] = new_val
                return True
    return False
 
 


# ═══════════════════════════════════════════════════════════════════════════════
#  SHARED PARSING HELPERS (survey workbooks / wide statistical tables)
# ═══════════════════════════════════════════════════════════════════════════════

def _num(x):
    """Parse a numeric cell that may contain commas, note codes like '[t]', or be numeric."""
    if x is None:
        return None
    if isinstance(x, (int, float)):
        return None if (isinstance(x, float) and x != x) else float(x)  # reject NaN
    s = re.sub(r"\[[a-z]+\]", "", str(x)).replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        return None


def _fmt(v, dp=1):
    """Format a float to dp decimal places, stripping trailing zeros sensibly."""
    if v is None:
        return None
    if dp == 0:
        return str(int(round(v)))
    return f"{round(v, dp):g}"


def _latest(value, period, **extra):
    d = {"value": value, "period": period}
    d.update(extra)
    return d


def _find_row(df, pattern, col=1):
    """Index of the first row whose `col` matches a regex (case-insensitive)."""
    pat = re.compile(pattern, re.I)
    for i, v in df[col].items():
        if pd.notna(v) and pat.search(str(v)):
            return i
    return None


def _last_num_in_row(df, row_i, start_col=2, end_col=None):
    """Rightmost numeric value in a row (and its column index)."""
    cols = range(start_col, end_col if end_col is not None else df.shape[1])
    best = (None, None)
    for c in cols:
        v = _num(df.iat[row_i, c]) if c < df.shape[1] else None
        if v is not None:
            best = (v, c)
    return best


# ═══════════════════════════════════════════════════════════════════════════════
#  FRS — FAMILY RESOURCES SURVEY (chapter 3 tenure)
# ═══════════════════════════════════════════════════════════════════════════════

def fetch_frs_tenure() -> dict | None:
    """
    Download ch3_tenure.xlsx from the latest FRS publication and parse:
      tenure_pct        — Table 3.1, PRS row, UK column (%)
      length_residency  — Table 3.4, PRS (all households block), <12 months (%)
      tenure_by_age     — Table 3.7, LATEST year block, 25-34 row, PRS column (%)
      median_rent       — Table 3.8, United Kingdom row, Private rented column (£/wk)
    """
    coll = "https://www.gov.uk/government/collections/family-resources-survey--2"
    try:
        r = SESSION.get(coll, timeout=30)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "html.parser")
        pub_url, year_label = None, None
        for a in soup.find_all("a", href=re.compile(r"family-resources-survey-financial-year-(\d{4})-to-(\d{4})")):
            m = re.search(r"financial-year-(\d{4})-to-(\d{4})", a["href"])
            href = a["href"]
            pub_url = href if href.startswith("http") else "https://www.gov.uk" + href
            year_label = f"{m.group(1)}/{m.group(2)[2:]}"
            break
        if not pub_url:
            warn("FRS: no publication found")
            return None
        log(f"  FRS: {pub_url} ({year_label})")

        pr = SESSION.get(pub_url, timeout=30)
        pr.raise_for_status()
        psoup = BeautifulSoup(pr.text, "html.parser")
        xlsx_url = None
        for a in psoup.find_all("a", href=re.compile(r"ch3_tenure\.xlsx$")):
            xlsx_url = a["href"]
            break
        if not xlsx_url:
            warn("FRS: ch3_tenure.xlsx not found on publication page")
            return None

        xr = SESSION.get(xlsx_url, timeout=60)
        xr.raise_for_status()
        wb = openpyxl.load_workbook(io.BytesIO(xr.content), read_only=True, data_only=True)
        out = {}

        # ── Table 3.1: PRS row, UK column ──
        rows = list(wb["3_1"].iter_rows(values_only=True))
        hdr_i = next(i for i, r in enumerate(rows) if r and str(r[0]).strip() == "Tenure")
        uk_col = next(j for j, c in enumerate(rows[hdr_i]) if c and "United Kingdom" in str(c))
        for r in rows[hdr_i + 1:]:
            if r and r[0] and "private rented" in str(r[0]).lower():
                v = _num(r[uk_col])
                if v is not None:
                    out["tenure_pct"] = _latest(_fmt(v, 0), year_label)
                break

        # ── Table 3.4: first PRS row (all-households block), 'Less than 12 months' col ──
        rows = list(wb["3_4"].iter_rows(values_only=True))
        hdr_i = next(i for i, r in enumerate(rows) if r and str(r[0]).strip() == "Tenure")
        lt12_col = next(j for j, c in enumerate(rows[hdr_i]) if c and "less than 12" in str(c).lower())
        for r in rows[hdr_i + 1:]:
            if r and r[0] and "private rented" in str(r[0]).lower():
                v = _num(r[lt12_col])
                if v is not None:
                    out["length_residency"] = _latest(_fmt(v, 0), year_label)
                break  # first occurrence = all households block

        # ── Table 3.7: LATEST year block, 25-34 row, PRS column ──
        rows = list(wb["3_7"].iter_rows(values_only=True))
        hdr_i = next(i for i, r in enumerate(rows) if r and str(r[0]).strip() == "Age")
        prs_col = next(j for j, c in enumerate(rows[hdr_i]) if c and "private rented" in str(c).lower())
        # year block markers look like '2014/15', '2024/25' in col 0; take rows after the LAST marker
        last_block = max(i for i, r in enumerate(rows) if r and re.match(r"^\d{4}/\d{2}$", str(r[0]).strip() if r[0] else ""))
        for r in rows[last_block + 1:]:
            if r and r[0] and re.match(r"^25\s*to\s*34", str(r[0]).strip()):
                v = _num(r[prs_col])
                if v is not None:
                    out["tenure_by_age"] = _latest(_fmt(v, 0), year_label)
                break

        # ── Table 3.8: United Kingdom row, Private rented column ──
        rows = list(wb["3_8"].iter_rows(values_only=True))
        hdr_i = next(i for i, r in enumerate(rows) if r and r[0] and "Region" in str(r[0]))
        prs_col = next(j for j, c in enumerate(rows[hdr_i]) if c and "private rented" in str(c).lower())
        for r in rows[hdr_i + 1:]:
            if r and r[0] and str(r[0]).strip() == "United Kingdom":
                v = _num(r[prs_col])
                if v is not None:
                    out["median_rent"] = _latest(_fmt(v, 0), year_label)
                break

        return out or None
    except Exception as e:
        warn(f"FRS tenure: {e}")
        return None


# ═══════════════════════════════════════════════════════════════════════════════
#  EHS — ENGLISH HOUSING SURVEY (headline annex tables)
# ═══════════════════════════════════════════════════════════════════════════════

EHS_COLLECTIONS = [
    ("demographics", r"english-housing-survey-(\d{4})-to-(\d{4})-headline-findings-on-demographics"),
    ("quality",      r"english-housing-survey-(\d{4})-to-(\d{4})-headline-findings-on-housing-quality"),
]


def _ehs_find_annex_files() -> tuple[dict, str] | tuple[None, None]:
    """Discover the four annex ODS files from the two EHS headline collections.
    Returns ({key: url}, '2024-25')."""
    base = "https://www.gov.uk/government/collections/english-housing-survey"
    r = SESSION.get(base, timeout=30)
    r.raise_for_status()
    soup = BeautifulSoup(r.text, "html.parser")

    files, year_label = {}, None
    for key, pattern in EHS_COLLECTIONS:
        coll_href = None
        for a in soup.find_all("a", href=re.compile(pattern)):
            coll_href = a["href"]
            m = re.search(pattern, coll_href)
            year_label = f"{m.group(1)}-{m.group(2)[2:]}"
            break
        if not coll_href:
            warn(f"EHS: collection not found for {key}")
            continue
        if not coll_href.startswith("http"):
            coll_href = "https://www.gov.uk" + coll_href

        cr = SESSION.get(coll_href, timeout=30)
        cr.raise_for_status()
        csoup = BeautifulSoup(cr.text, "html.parser")
        annex_href = None
        for a in csoup.find_all("a", href=re.compile(r"/government/statistics/annex-tables-for-")):
            annex_href = a["href"]
            break
        if not annex_href:
            warn(f"EHS: annex page not found for {key}")
            continue
        if not annex_href.startswith("http"):
            annex_href = "https://www.gov.uk" + annex_href

        ar = SESSION.get(annex_href, timeout=30)
        ar.raise_for_status()
        asoup = BeautifulSoup(ar.text, "html.parser")
        for a in asoup.find_all("a", href=re.compile(r"\.ods$")):
            href = a["href"]
            fname = href.split("/")[-1].lower()
            if "profile_of_households" in fname:
                files["profile"] = href
            elif "costs_and_affordability" in fname:
                files["costs"] = href
            elif "housing_quality" in fname:
                files["quality"] = href
            elif "energy_efficiency" in fname:
                files["energy"] = href
    return (files, year_label) if files else (None, None)


def _ods(url) -> dict[str, pd.DataFrame]:
    r = SESSION.get(url, timeout=120)
    r.raise_for_status()
    xl = pd.ExcelFile(io.BytesIO(r.content), engine="odf")
    sheets = {}
    for s in xl.sheet_names:
        if re.match(r"^AT\d+_\d+$", s):
            sheets[s] = pd.read_excel(xl, sheet_name=s, header=None)
    return sheets


def fetch_ehs_annex() -> dict | None:
    """Parse all 13 EHS dashboard metrics from the four 2024-25 annex ODS files."""
    files, year_label = _ehs_find_annex_files()
    if not files:
        return None
    log(f"  EHS: {year_label}, files found: {sorted(files)}")
    out = {}
    cal_year = year_label.split("-")[0] if year_label else None
    # physical-survey tables are labelled by calendar year (e.g. 2024)
    phys_year = str(int(cal_year)) if cal_year else None

    # ── Profile of households and dwellings ──
    if "profile" in files:
        sh = _ods(files["profile"])

        # AT1_1: trends in tenure — last data row = latest year
        df = sh.get("AT1_1")
        if df is not None:
            # year rows look like '2024-25' in col 1
            yr_rows = [i for i, v in df[1].items() if pd.notna(v) and re.match(r"^\d{4}-\d{2}$", str(v).strip())]
            if yr_rows:
                i = yr_rows[-1]
                period = str(df.iat[i, 1]).strip()
                owners_all = _num(df.iat[i, 4])   # all owner occupiers
                prs = _num(df.iat[i, 5])          # private renters (000s)
                social = _num(df.iat[i, 8])       # all social renters
                if prs is not None:
                    out["households_tenure"] = _latest(_fmt(prs, 0), period)
                    if owners_all and social:
                        pct = prs / (owners_all + prs + social) * 100
                        out["households_tenure_pct"] = _latest(_fmt(pct, 1), period)

        # AT1_2: tenure by region — London block, PRS share of all households, latest col
        df = sh.get("AT1_2")
        if df is not None:
            lon_i = _find_row(df, r"^London$")
            if lon_i is not None:
                # rows below London: owner occupiers / own outright / buying / private renters / social renters
                block = {}
                for i in range(lon_i + 1, min(lon_i + 9, len(df))):
                    label = str(df.iat[i, 1]).strip().lower() if pd.notna(df.iat[i, 1]) else ""
                    if label in ("owner occupiers", "private renters", "social renters", "all households", "all tenures"):
                        v, c = _last_num_in_row(df, i)
                        if v is not None:
                            block[label] = v
                prs = block.get("private renters")
                total = block.get("all households") or block.get("all tenures") or \
                    (sum(v for k, v in block.items() if k in ("owner occupiers", "private renters", "social renters")) or None)
                if prs and total:
                    out["tenure_by_region"] = _latest(
                        _fmt(prs / total * 100, 1), year_label,
                        notes_hint="London PRS share")

        # AT1_9: vacant dwellings by tenure — England total dwellings (occupied + vacant)
        df = sh.get("AT1_9")
        if df is not None:
            # year header row: cells like 2019.0 / 2024 in the counts block
            yr_hdr = None
            for i in range(min(8, len(df))):
                yrs = [(c, _num(df.iat[i, c])) for c in range(2, df.shape[1])]
                yrs = [(c, v) for c, v in yrs if v and 2000 <= v <= 2100]
                if len(yrs) >= 3:
                    yr_hdr = yrs
                    break
            if yr_hdr:
                # counts block = first contiguous run of year columns
                cols = [c for c, _v in yr_hdr]
                count_cols = [cols[0]]
                for c in cols[1:]:
                    if c == count_cols[-1] + 1:
                        count_cols.append(c)
                    else:
                        break
                latest_col = count_cols[-1]
                latest_yr = str(int(dict(yr_hdr)[latest_col]))
                total = 0.0
                for i, v in df[1].items():
                    if pd.notna(v) and re.match(r"^all (owner|private|social)", str(v).strip(), re.I):
                        n = _num(df.iat[i, latest_col])
                        if n:
                            total += n
                if total:
                    out["dwellings_vacant"] = _latest(_fmt(total, 0), latest_yr)

        # AT1_6: stock profile — dwelling age (pre-1919 PRS share), rural/urban split
        df = sh.get("AT1_6")
        if df is not None:
            age_i = _find_row(df, r"^dwelling age$")
            if age_i is not None:
                ages = {}
                for i in range(age_i + 1, min(age_i + 12, len(df))):
                    label = str(df.iat[i, 1]).strip() if pd.notna(df.iat[i, 1]) else ""
                    if not label:
                        break
                    v = _num(df.iat[i, 3])  # private rented col
                    if v is not None:
                        ages[label] = v
                if ages:
                    total = sum(ages.values())
                    pre1919 = ages.get("pre-1919")
                    if pre1919 and total:
                        out["age_of_property"] = _latest(
                            _fmt(pre1919 / total * 100, 1), phys_year or year_label,
                            notes_hint="PRS dwellings built pre-1919")
            # 'type of area' block → rural share of PRS stock
            area_i = _find_row(df, r"^type of area$")
            all_i = None
            if area_i is not None:
                for i, v in df[1].items():
                    if i > area_i and pd.notna(v) and str(v).strip().lower() == "all dwellings":
                        all_i = i
                        break
            if area_i is not None and all_i is not None:
                rural = 0.0
                for i in range(area_i + 1, min(area_i + 10, len(df))):
                    label = str(df.iat[i, 1]).strip().lower() if pd.notna(df.iat[i, 1]) else ""
                    if not label:
                        continue
                    if label.startswith(("deprived", "occupancy", "floor")):
                        break
                    if "rural" in label or "village" in label:
                        v = _num(df.iat[i, 3])  # private rented col
                        if v is not None:
                            rural += v
                all_prs = _num(df.iat[all_i, 3])
                if rural and all_prs:
                    out["rural_urban_tenure"] = _latest(
                        _fmt(rural / all_prs * 100, 1), phys_year or year_label,
                        notes_hint="PRS dwellings in rural areas/villages")

    # ── Housing costs and affordability ──
    if "costs" in files:
        sh = _ods(files["costs"])

        # AT2_4: mean & median weekly rents — England block (second geography), private renters
        df = sh.get("AT2_4")
        if df is not None:
            # geography markers in col 1: 'London', 'England' (or 'England (excluding London)'... )
            geo_rows = [(i, str(df.iat[i, 1]).strip()) for i, v in df[1].items()
                        if pd.notna(v) and re.match(r"^(London|England)", str(v).strip())
                        and not re.match(r"^\d", str(v).strip())]
            eng_i = None
            for i, label in geo_rows:
                if label.lower().startswith("england"):
                    eng_i = i
            if eng_i is None and geo_rows:
                eng_i = geo_rows[-1][0]
            # find latest year row within England block
            yr_rows = [i for i, v in df[1].items()
                       if i > eng_i and pd.notna(v) and re.match(r"^\d{4}-\d{2}$", str(v).strip())]
            if yr_rows:
                i = yr_rows[-1]
                period = str(df.iat[i, 1]).strip()
                mean_prs = _num(df.iat[i, 2])    # private renters mean
                median_prs = _num(df.iat[i, 8])  # private renters median (right block)
                if mean_prs is not None:
                    out["mean_weekly_rents"] = _latest(_fmt(mean_prs, 0), period)
                if median_prs is not None:
                    out["median_weekly_rents"] = _latest(_fmt(median_prs, 0), period)

        # AT2_5: rent as % of income — private renters, latest year row whose value
        # is a plausible percentage (later blocks in the sheet hold sample sizes)
        df = sh.get("AT2_5")
        if df is not None:
            yr_rows = [i for i, v in df[1].items() if pd.notna(v) and re.match(r"^\d{4}-\d{2}$", str(v).strip())]
            best = None
            for i in yr_rows:
                v = _num(df.iat[i, 3])  # private renters column
                if v is not None and 0 < v <= 100:
                    best = (i, v)
            if best:
                i, v = best
                out["rent_as_pct_income"] = _latest(
                    _fmt(v, 1), str(df.iat[i, 1]).strip(),
                    notes_hint="household income basis, incl. housing support")

    # ── Housing quality ──
    if "quality" in files:
        sh = _ods(files["quality"])

        def prs_pct_from_year_cols(df):
            """Tables AT1_4/AT1_6: year header row + tenure rows, repeated in
            counts / percentage / sample-size blocks. Return (pct, year) from the
            'private rented' row whose values are plausible percentages."""
            # year header: row with several 4-digit year cells
            yr_map = None
            for i in range(min(8, len(df))):
                yrs = {c: _num(df.iat[i, c]) for c in range(2, df.shape[1])}
                yrs = {c: int(v) for c, v in yrs.items() if v and 2000 <= v <= 2100}
                if len(yrs) >= 3:
                    yr_map = yrs
                    break
            if not yr_map:
                return None, None
            for i, v in df[1].items():
                if pd.notna(v) and str(v).strip().lower() == "private rented":
                    # rightmost numeric in year columns
                    best = None
                    for c in sorted(yr_map):
                        n = _num(df.iat[i, c])
                        if n is not None:
                            best = (n, yr_map[c])
                    if best and 0 < best[0] <= 100:
                        return best
            return None, None

        for sheet, key in (("AT1_4", "non_decent_homes"), ("AT1_6", "hhsrs_cat1")):
            df = sh.get(sheet)
            if df is None:
                continue
            v, yr = prs_pct_from_year_cols(df)
            if v is not None:
                out[key] = _latest(_fmt(v, 1), str(yr))

        # AT1_10: damp problems by tenure — percentage block, 'any damp problem' col
        df = sh.get("AT1_10")
        if df is not None:
            damp_col = None
            for i in range(min(8, len(df))):
                for c in range(df.shape[1]):
                    v = df.iat[i, c]
                    if pd.notna(v) and "any damp" in str(v).lower():
                        damp_col = c
                        break
                if damp_col:
                    break
            if damp_col is not None:
                for i, v in df[1].items():
                    if pd.notna(v) and str(v).strip().lower() == "private rented":
                        n = _num(df.iat[i, damp_col])
                        if n is not None and 0 < n <= 100:
                            out["damp_problems"] = _latest(_fmt(n, 1), phys_year or year_label)
                            # keep iterating: percentage block comes after counts block

    # ── Energy efficiency ──
    if "energy" in files:
        sh = _ods(files["energy"])
        # AT2_2: EER bands by tenure — PRS block, latest year row: A-to-C / all dwellings
        df = sh.get("AT2_2")
        if df is not None:
            atoc_col = all_col = None
            for i in range(min(8, len(df))):
                for c in range(df.shape[1]):
                    v = str(df.iat[i, c]).strip().lower() if pd.notna(df.iat[i, c]) else ""
                    if v in ("a to c", "a-c"):
                        atoc_col = c
                    elif v == "all dwellings" and atoc_col is not None:
                        all_col = c
                if atoc_col is not None and all_col is not None:
                    break
            pr_i = None
            for i, v in df[1].items():
                if pd.notna(v) and str(v).strip().lower() == "private rented":
                    pr_i = i
                    break
            if None not in (atoc_col, all_col, pr_i):
                latest = None
                for i in range(pr_i + 1, min(pr_i + 6, len(df))):
                    label = str(df.iat[i, 1]).strip() if pd.notna(df.iat[i, 1]) else ""
                    if label and not re.match(r"^\d{4}$", label):
                        break  # next tenure block (e.g. 'all private sector')
                    if re.match(r"^\d{4}$", label):
                        atoc, alln = _num(df.iat[i, atoc_col]), _num(df.iat[i, all_col])
                        if atoc and alln:
                            latest = (atoc / alln * 100, label)
                if latest:
                    out["energy_efficiency"] = _latest(_fmt(latest[0], 1), latest[1])

    return out or None


# ═══════════════════════════════════════════════════════════════════════════════
#  MoJ — POSSESSION TIMELINESS (Tables ODS, Data_6)
# ═══════════════════════════════════════════════════════════════════════════════

def fetch_moj_timeliness() -> dict | None:
    """Mean/median weeks from landlord possession claim to order, latest quarter.
    Source: Data_6 sheet of the Mortgage & Landlord Possession Tables ODS."""
    coll = "https://www.gov.uk/government/collections/mortgage-and-landlord-possession-statistics"
    try:
        r = SESSION.get(coll, timeout=30)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "html.parser")
        pub_url = None
        for a in soup.find_all("a", href=re.compile(
                r"/government/statistics/mortgage-and-landlord-possession-statistics-(?!earlier|guide)")):
            href = a["href"]
            pub_url = href if href.startswith("http") else "https://www.gov.uk" + href
            break
        if not pub_url:
            return None
        pr = SESSION.get(pub_url, timeout=30)
        pr.raise_for_status()
        psoup = BeautifulSoup(pr.text, "html.parser")
        ods_url = None
        for a in psoup.find_all("a", href=re.compile(r"Possession_Tables.*\.ods$", re.I)):
            if "accessible" not in a["href"].lower():
                ods_url = a["href"]
                break
        if not ods_url:
            return None
        log(f"  MoJ timeliness: {ods_url.split('/')[-1]}")
        orq = SESSION.get(ods_url, timeout=120)
        orq.raise_for_status()
        df = pd.read_excel(io.BytesIO(orq.content), sheet_name="Data_6", engine="odf", header=0)
        # LOOKUP col: 'YYYY|Qn|LandlordType|OrderType'
        lookup_col = df.columns[0]
        recs = {}
        for _, row in df.iterrows():
            key = str(row[lookup_col])
            m = re.match(r"^(\d{4})\|(Q\d)\|([^|]+)\|All", key)
            if not m:
                continue
            year, q, ltype = m.group(1), m.group(2), m.group(3)
            recs.setdefault((year, q), {})[ltype] = row
        if not recs:
            return None
        latest = max(recs.keys(), key=lambda k: (k[0], k[1]))
        period = f"{latest[0]} {latest[1]}"
        block = recs[latest]
        out = {}
        if "All" in block:
            mean_v = _num(block["All"].get("MeanTime_Order"))
            med_v = _num(block["All"].get("MedTime_Order"))
            if mean_v is not None:
                out["mean_time_all"] = _latest(_fmt(mean_v, 1), period)
            if med_v is not None:
                out["median_time_all"] = _latest(_fmt(med_v, 1), period)
        if "Private_Landlord" in block:
            v = _num(block["Private_Landlord"].get("MeanTime_Order"))
            if v is not None:
                out["mean_time_prs"] = _latest(_fmt(v, 1), period)
        return out or None
    except Exception as e:
        warn(f"MoJ timeliness: {e}")
        return None


# ═══════════════════════════════════════════════════════════════════════════════
#  MLAR — FCA/BoE Mortgage Lenders & Administrators Return (long-run xlsx)
# ═══════════════════════════════════════════════════════════════════════════════

MLAR_DETAILED_URL = "https://www.fca.org.uk/publication/data/mlar-statistics-detailed-long-run.xlsx"


def fetch_mlar() -> dict | None:
    """Gross advances (Table 1.21, regulated + non-regulated, £bn) and
    buy-to-let share of all residential gross advances (Table 1.33 §C line 4, %)."""
    try:
        r = SESSION.get(MLAR_DETAILED_URL, timeout=120)
        r.raise_for_status()
        wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
        out = {}

        def grid(sheet):
            return [list(row) for row in wb[sheet].iter_rows(values_only=True)]

        def period_cols(rows):
            """Map column index → 'YYYY Qn' using the year row (sparse) and quarter row."""
            year_i = quarter_i = None
            for i, row in enumerate(rows[:20]):
                vals = [str(c).strip() for c in row if c is not None]
                if vals and all(re.match(r"^\d{4}(\.0)?$", v) for v in vals) and len(vals) > 3:
                    year_i = i
                if vals and all(re.match(r"^Q[1-4]$", v) for v in vals) and len(vals) > 3:
                    quarter_i = i
                    break
            if year_i is None or quarter_i is None:
                return {}
            periods, cur_year = {}, None
            for c in range(len(rows[quarter_i])):
                y = rows[year_i][c] if c < len(rows[year_i]) else None
                if y is not None and re.match(r"^\d{4}", str(y)):
                    cur_year = str(int(float(str(y))))
                q = rows[quarter_i][c]
                if q is not None and re.match(r"^Q[1-4]$", str(q).strip()) and cur_year:
                    periods[c] = f"{cur_year} {str(q).strip()}"
            return periods

        def section_line(rows, section_letter, line_no, label_hint):
            """Find data row for sub-table line `line_no` within section `section_letter`."""
            sec_i = None
            for i, row in enumerate(rows):
                a = str(row[0]).strip() if row[0] is not None else ""
                if a == section_letter:
                    sec_i = i
                elif sec_i is not None and re.match(r"^[A-Z]$", a) and i > sec_i:
                    break  # next section
                if sec_i is not None and a == str(line_no):
                    label = " ".join(str(x).strip().lower() for x in row[1:5] if x is not None)
                    if label_hint in label:
                        return i
            return None

        # ── 1.21 gross advances: A1 + B1 ──
        rows121 = grid("1.21")
        p121 = period_cols(rows121)
        ia = section_line(rows121, "A", 1, "gross advances")
        ib = section_line(rows121, "B", 1, "gross advances")
        if ia is not None and ib is not None and p121:
            c = max(c for c in p121 if _num(rows121[ia][c]) is not None)
            total_m = (_num(rows121[ia][c]) or 0) + (_num(rows121[ib][c]) or 0)
            if total_m:
                out["gross_advances"] = _latest(_fmt(total_m / 1000, 1), p121[c])

        # ── 1.33 BTL share: section C line 4 ──
        rows133 = grid("1.33")
        p133 = period_cols(rows133)
        ic = section_line(rows133, "C", 4, "buy to let")
        if ic is not None and p133:
            c = max(c for c in p133 if _num(rows133[ic][c]) is not None)
            v = _num(rows133[ic][c])
            if v is not None and 0 < v < 100:
                out["btl_proportion"] = _latest(_fmt(v, 1), p133[c])
        return out or None
    except Exception as e:
        warn(f"MLAR: {e}")
        return None


# ═══════════════════════════════════════════════════════════════════════════════
#  SECTION RUNNERS
# ═══════════════════════════════════════════════════════════════════════════════
 

def update_macro():
    log("=== Updating Macro-Economic data ===")
    path = DATA_DIR / "macro.json"
    data = load_json(path)
    ds = data["datasets"]
    changed = False
 
    # Auto-remove any datasets marked as ceased (keeps GitHub JSON clean)
    original_len = len(ds)
    ds[:] = [d for d in ds if d.get("status") != "ceased"]
    if len(ds) < original_len:
        log(f"  Removed {original_len - len(ds)} ceased dataset(s)")
        changed = True
 
    ons_cdid_map = {
        "cpi_annual_rate":      "D7G7",
        "cpi_index":            "D7BT",
        "cpih_annual_rate":     "L55O",
        "cpih_index":           "L522",
        "gdp_qoq":              "IHYQ",
        "gdp_yoy":              "IHYR",
        "awe_total_pay":        "K54U",
        "real_earnings_index":  "A2FD",
        "real_earnings_growth": "A3WV",
    }
 
    for dataset_id, cdid in ons_cdid_map.items():
        log(f"  ONS {cdid} → {dataset_id}")
        result = fetch_ons_timeseries(cdid)
        if update_dataset(ds, dataset_id, result):
            log(f"    Updated: {result}")
            changed = True
        else:
            log(f"    No change (fetched: {result})")
        time.sleep(0.3)
 
    # PIPR — Price Index of Private Rents (UK annual rate + index)
    # Downloaded from the ONS dataset xlsx (the bulletin scrape was unreliable).
    log("  PIPR (UK): annual rate + index")
    pipr = fetch_ons_pipr_uk()
    if pipr:
        if update_dataset(ds, "pipr_annual_rate_uk", pipr.get("annual_rate")):
            log(f"    pipr_annual_rate_uk: {pipr['annual_rate']}")
            changed = True
        if update_dataset(ds, "pipr_index_uk", pipr.get("index")):
            log(f"    pipr_index_uk: {pipr['index']}")
            changed = True
    else:
        log("    No PIPR data retrieved")
    time.sleep(0.3)
 
    # Bank of England base rate
    log("  BoE base rate (IUMABEDR)")
    boe_rate = fetch_boe_series("IUMABEDR", "base rate")
    if update_dataset(ds, "boe_base_rate", boe_rate):
        log(f"    Updated: {boe_rate}")
        changed = True
    else:
        log(f"    No change (fetched: {boe_rate})")
 
    # MLAR — residential gross advances + buy-to-let share.
    # The BoE IADB blocks the MLAR series, so we parse the FCA's co-published
    # long-run detailed tables instead (Table 1.21 §A+§B and Table 1.33 §C).
    log("  MLAR (FCA long-run tables): gross advances + BTL share")
    mlar = fetch_mlar()
    if mlar:
        if update_dataset(ds, "boe_gross_advances", mlar.get("gross_advances")):
            log(f"    boe_gross_advances: {mlar['gross_advances']}")
            changed = True
        if update_dataset(ds, "boe_btl_proportion", mlar.get("btl_proportion")):
            log(f"    boe_btl_proportion: {mlar['btl_proportion']}")
            changed = True
    else:
        log("    No MLAR data retrieved")
 
    if changed:
        data["last_fetched"] = now_iso()
        save_json(path, data)
        log("  macro.json saved ✓")
    else:
        log("  No changes to macro.json")
 
 

def update_housing_tenure():
    log("=== Updating Housing Trends (FRS) ===")
    path = DATA_DIR / "housing_tenure.json"
    data = load_json(path)
    ds = data["datasets"]

    frs = fetch_frs_tenure()
    if not frs:
        log("  No FRS data retrieved")
        return

    changed = False
    mapping = {
        "tenure_pct":       "frs_tenure_pct",
        "length_residency": "frs_length_of_residency",
        "tenure_by_age":    "frs_tenure_by_age",
        "median_rent":      "frs_median_rent_mortgage",
    }
    for key, ds_id in mapping.items():
        if key in frs and update_dataset(ds, ds_id, frs[key]):
            log(f"  {ds_id}: {frs[key]}")
            changed = True

    if changed:
        data["last_fetched"] = now_iso()
        save_json(path, data)
        log("  housing_tenure.json saved ✓")
    else:
        log("  No changes to housing_tenure.json")


def update_dwellings():
    log("=== Updating Dwellings data ===")
    path = DATA_DIR / "dwellings.json"
    data = load_json(path)
    ds = data["datasets"]
    changed = False
 
    log("  MHCLG Table 213 (England completions)")
    mhclg = fetch_mhclg_table213()
    if mhclg:
        for key, dst_id in [
            ("private", "mhclg_completions_private"),
            ("ha",      "mhclg_completions_ha"),
            ("la",      "mhclg_completions_la"),
            ("total",   "mhclg_completions_total"),
        ]:
            if key in mhclg and update_dataset(ds, dst_id, mhclg[key]):
                log(f"    {dst_id}: {mhclg[key]}")
                changed = True
 
    log("  Stats Wales (Wales completions)")
    wales = fetch_stats_wales_completions()
    if wales:
        for key, dst_id in [
            ("total",   "stats_wales_completions_total"),
            ("private", "stats_wales_completions_private"),
        ]:
            if key in wales and update_dataset(ds, dst_id, wales[key]):
                log(f"    {dst_id}: {wales[key]}")
                changed = True
 
    if changed:
        data["last_fetched"] = now_iso()
        save_json(path, data)
        log("  dwellings.json saved ✓")
    else:
        log("  No changes to dwellings.json")
 
 

def update_possession():
    log("=== Updating Claims & Bailiffs (MoJ) ===")
    path = DATA_DIR / "possession.json"
    data = load_json(path)
    ds = data["datasets"]
 
    moj = fetch_moj_possession()
    if not moj:
        log("  No MoJ data retrieved")
        return
 
    changed = False
    # Timeliness (mean/median weeks claim→order) lives in the Tables ODS,
    # not the CSV zip — fetch and merge it in.
    timeliness = fetch_moj_timeliness()
    if timeliness:
        moj.update(timeliness)

    mapping = {
        "claims_issued":          "mlp_claims_issued",
        "repossessions_bailiffs": "mlp_repossessions_bailiffs",
        "claims_prs":             "mlp_claims_prs",
        "claims_accelerated":     "mlp_claims_accelerated",
        "mean_time_all":          "mlp_mean_time_all",
        "mean_time_prs":          "mlp_mean_time_prs",
        "median_time_all":        "mlp_median_time_all",
    }
    for moj_key, ds_id in mapping.items():
        if moj_key in moj and update_dataset(ds, ds_id, moj[moj_key]):
            log(f"  {ds_id}: {moj[moj_key]}")
            changed = True
 
    if changed:
        data["last_fetched"] = now_iso()
        save_json(path, data)
        log("  possession.json saved ✓")
    else:
        log("  No changes to possession.json")
 
 

def update_ehs():
    log("=== Updating English Housing Survey ===")
    path = DATA_DIR / "ehs.json"
    data = load_json(path)
    ds = data["datasets"]

    ehs = fetch_ehs_annex()
    if not ehs:
        log("  No EHS data retrieved")
        return

    changed = False
    mapping = {
        "dwellings_vacant":      "ehs_dwellings_vacant",
        "households_tenure":     "ehs_households_tenure",
        "households_tenure_pct": "ehs_households_tenure_pct",
        "tenure_by_region":      "ehs_tenure_by_region",
        "mean_weekly_rents":     "ehs_mean_weekly_rents",
        "median_weekly_rents":   "ehs_median_weekly_rents",
        "rent_as_pct_income":    "ehs_rent_as_pct_income",
        "age_of_property":       "ehs_age_of_property",
        "rural_urban_tenure":    "ehs_rural_urban_tenure",
        "non_decent_homes":      "ehs_non_decent_homes",
        "hhsrs_cat1":            "ehs_hhsrs_cat1",
        "damp_problems":         "ehs_damp_problems",
        "energy_efficiency":     "ehs_energy_efficiency",
    }
    for key, ds_id in mapping.items():
        if key in ehs and update_dataset(ds, ds_id, ehs[key]):
            log(f"  {ds_id}: {ehs[key]}")
            changed = True

    if changed:
        data["last_fetched"] = now_iso()
        save_json(path, data)
        log("  ehs.json saved ✓")
    else:
        log("  No changes to ehs.json")


def main():
    log(f"NRLA Data Dashboard updater — {datetime.now().strftime('%Y-%m-%d %H:%M UTC')}")
    log(f"Data directory: {DATA_DIR}")
 
    if not DATA_DIR.exists():
        error(f"Data directory not found: {DATA_DIR}")
        sys.exit(1)
 
    errors = []
    for label, fn in [
        ("Macro-Economic",  update_macro),
        ("Housing Tenure",  update_housing_tenure),
        ("Dwellings",       update_dwellings),
        ("Claims/Bailiffs", update_possession),
        ("EHS",             update_ehs),
    ]:
        try:
            fn()
        except Exception as e:
            msg = f"{label}: unexpected error — {e}"
            error(msg)
            errors.append(msg)
 
    if errors:
        log(f"\nCompleted with {len(errors)} error(s):")
        for e in errors:
            log(f"  ✖ {e}")
        sys.exit(1)
    else:
        log("\nAll sections updated successfully ✓")
 
 
if __name__ == "__main__":
    main()
 
